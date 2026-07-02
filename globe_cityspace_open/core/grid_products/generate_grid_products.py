from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook


ROWS = 8
COLS = 16
DPI = 254
PX_PER_CM = 100


def load_contract(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def build_matrix(cells: list[dict]) -> list[list[dict]]:
    matrix = [[None for _ in range(COLS)] for _ in range(ROWS)]

    for cell in cells:
        r = int(cell["row"]) - 1
        c = int(cell["col"]) - 1
        matrix[r][c] = cell

    return matrix


def write_csv(path: Path, cells: list[dict], cell_size_cm: int) -> None:
    fields = [
        "cell_id",
        "pin_id",
        "row",
        "col",
        "x_prime_cm",
        "y_prime_cm",
        "x_prime_center_cm",
        "y_prime_center_cm",
        "cell_size_cm",
        "terrain_height_m",
        "building_height_m",
        "total_height_m",
        "relative_height_m",
        "pin_height_cm",
        "gray_value",
    ]

    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for cell in cells:
            row = int(cell["row"])
            col = int(cell["col"])

            x0 = (col - 1) * cell_size_cm
            y0 = (row - 1) * cell_size_cm

            w.writerow({
                "cell_id": cell["cell_id"],
                "pin_id": cell["pin_id"],
                "row": row,
                "col": col,
                "x_prime_cm": x0,
                "y_prime_cm": y0,
                "x_prime_center_cm": x0 + cell_size_cm / 2,
                "y_prime_center_cm": y0 + cell_size_cm / 2,
                "cell_size_cm": cell_size_cm,
                "terrain_height_m": cell["terrain_height_m"],
                "building_height_m": cell["building_height_m"],
                "total_height_m": cell["total_height_m"],
                "relative_height_m": cell["relative_height_m"],
                "pin_height_cm": cell["pin_height_cm"],
                "gray_value": cell["gray_value"],
            })


def write_xlsx(path: Path, cells: list[dict], matrix: list[list[dict]], cell_size_cm: int) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "cells"

    headers = [
        "cell_id", "pin_id", "row", "col",
        "x_prime_cm", "y_prime_cm",
        "x_prime_center_cm", "y_prime_center_cm",
        "cell_size_cm",
        "terrain_height_m", "building_height_m",
        "total_height_m", "relative_height_m",
        "pin_height_cm", "gray_value"
    ]
    ws.append(headers)

    for cell in cells:
        row = int(cell["row"])
        col = int(cell["col"])
        x0 = (col - 1) * cell_size_cm
        y0 = (row - 1) * cell_size_cm

        ws.append([
            cell["cell_id"],
            cell["pin_id"],
            row,
            col,
            x0,
            y0,
            x0 + cell_size_cm / 2,
            y0 + cell_size_cm / 2,
            cell_size_cm,
            cell["terrain_height_m"],
            cell["building_height_m"],
            cell["total_height_m"],
            cell["relative_height_m"],
            cell["pin_height_cm"],
            cell["gray_value"],
        ])

    ws2 = wb.create_sheet("gray_matrix")
    for r in range(ROWS):
        ws2.append([matrix[r][c]["gray_value"] for c in range(COLS)])

    ws3 = wb.create_sheet("cell_id_matrix")
    for r in range(ROWS):
        ws3.append([matrix[r][c]["cell_id"] for c in range(COLS)])

    wb.save(path)


def write_gray_images(out_prefix: Path, matrix: list[list[dict]], cell_size_cm: int) -> None:
    cell_px = cell_size_cm * PX_PER_CM
    width = COLS * cell_px
    height = ROWS * cell_px

    img = Image.new("L", (width, height), 0)
    draw = ImageDraw.Draw(img)

    for r in range(ROWS):
        for c in range(COLS):
            gray = int(matrix[r][c]["gray_value"])
            x0 = c * cell_px
            y0 = r * cell_px
            x1 = x0 + cell_px
            y1 = y0 + cell_px
            draw.rectangle([x0, y0, x1, y1], fill=gray)

    img.save(str(out_prefix) + ".bmp", dpi=(DPI, DPI))
    img.save(str(out_prefix) + ".png", dpi=(DPI, DPI))


def write_validation_png(path: Path, matrix: list[list[dict]], cell_size_cm: int) -> None:
    cell_px = cell_size_cm * PX_PER_CM
    width = COLS * cell_px
    height = ROWS * cell_px

    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.load_default()
    except Exception:
        font = None

    for r in range(ROWS):
        for c in range(COLS):
            cell = matrix[r][c]
            gray = int(cell["gray_value"])

            x0 = c * cell_px
            y0 = r * cell_px
            x1 = x0 + cell_px
            y1 = y0 + cell_px

            draw.rectangle([x0, y0, x1, y1], fill=(gray, gray, gray), outline=(255, 0, 0), width=2)

            text = f'{cell["cell_id"]}\nG={gray}\nZ={cell["pin_height_cm"]:.2f}cm'
            text_color = (0, 0, 0) if gray > 150 else (255, 255, 255)
            draw.multiline_text((x0 + 6, y0 + 6), text, fill=text_color, font=font)

    draw.text((10, 10), "(0,0,0) mesa = canto superior esquerdo | x' direita | y' baixo | z' cima", fill=(255, 0, 0), font=font)

    img.save(path, dpi=(DPI, DPI))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--height-contract", required=True)
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()

    height_contract = load_contract(Path(args.height_contract))
    cells = height_contract["cells"]
    matrix = build_matrix(cells)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    for cell_size_cm in [1, 2]:
        suffix = f"{cell_size_cm}cm"

        write_csv(out_dir / f"grid_gray_{suffix}.csv", cells, cell_size_cm)
        write_xlsx(out_dir / f"grid_gray_{suffix}.xlsx", cells, matrix, cell_size_cm)
        write_gray_images(out_dir / f"grid_gray_{suffix}", matrix, cell_size_cm)
        write_validation_png(out_dir / f"grid_gray_{suffix}_validation.png", matrix, cell_size_cm)

    manifest = {
        "contract_type": "grid_products",
        "contract_version": "G5.6",
        "source_height_contract": str(args.height_contract),
        "coordinate_system": {
            "origin": "upper_left",
            "origin_xyz_cm": [0, 0, 0],
            "x_prime": "right",
            "y_prime": "down",
            "z_prime": "up"
        },
        "grid": {
            "cols_x_prime": 16,
            "rows_y_prime": 8,
            "supported_cell_sizes_cm": [1, 2]
        },
        "products": [
            "grid_gray_1cm.bmp",
            "grid_gray_1cm.png",
            "grid_gray_1cm.csv",
            "grid_gray_1cm.xlsx",
            "grid_gray_1cm_validation.png",
            "grid_gray_2cm.bmp",
            "grid_gray_2cm.png",
            "grid_gray_2cm.csv",
            "grid_gray_2cm.xlsx",
            "grid_gray_2cm_validation.png"
        ]
    }

    (out_dir / "grid_products_manifest_G56.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )

    print("GRID PRODUCTS CREATED")
    print(out_dir)


if __name__ == "__main__":
    main()
